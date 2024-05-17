#For google colab
import os
from Crypto.Cipher import AES
from Crypto.Random import get_random_bytes
from openpyxl import Workbook, load_workbook
from Crypto.Hash import SHA256
import base64
import logging
from datetime import datetime
import qrcode
import time
from PIL import Image
import io
from IPython.display import display

# Set up basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def input_private_key():
    """Prompt user to input their own private key and convert it to a suitable AES key."""
    key_input = input("Enter your private key (any string, will be hashed to create a key): ")
    hasher = SHA256.new()
    hasher.update(key_input.encode('utf-8'))
    raw_key = hasher.digest()  # This will be a 32-byte key
    base64_key = base64.b64encode(raw_key).decode('utf-8')
    return raw_key, base64_key

def generate_private_key():
    """Generate a new private key and encode it in base64, ensuring it's 32 bytes (AES-256)."""
    private_key = get_random_bytes(32)  # 256-bit AES key
    encoded_key = base64.b64encode(private_key).decode('utf-8')
    return private_key, encoded_key

def save_data_to_excel(date_time, key, original_text, encrypted_text, qr_code_path):
    """Save the data to an Excel file."""
    filename = 'keys.xlsx'
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Date and Time', 'Key (Base64)', 'Original Text', 'Encrypted Text', 'QR Code'])
    else:
        ws = wb.active

    ws.append([date_time, key, original_text, encrypted_text, qr_code_path])
    wb.save(filename)
    logging.info("Data saved to keys.xlsx")

def encrypt_data(data, key):
    """Encrypt the data using AES encryption with a given key."""
    cipher = AES.new(key, AES.MODE_EAX)
    ciphertext, tag = cipher.encrypt_and_digest(data.encode('utf-8'))
    return base64.b64encode(ciphertext).decode(), base64.b64encode(cipher.nonce).decode(), base64.b64encode(tag).decode()

def decrypt_data(ciphertext, nonce, tag, key):
    """Decrypt the ciphertext using AES decryption with the given key."""
    try:
        ciphertext = base64.b64decode(ciphertext)
        nonce = base64.b64decode(nonce)
        tag = base64.b64decode(tag)
        cipher = AES.new(key, AES.MODE_EAX, nonce=nonce)
        plaintext = cipher.decrypt_and_verify(ciphertext, tag)
        return plaintext.decode('utf-8')
    except ValueError:
        logging.error("Decryption failed.")
        return "Decryption failed."

def assess_encryption_strength(ciphertext):
    """Assess the strength of the encryption based on ciphertext length."""
    length = len(ciphertext)
    if length < 64:
        return "Weak (less than 50%)"
    elif length < 128:
        return "Moderate (50% - 75%)"
    else:
        return "Strong (greater than 75%)"

def generate_qr_code(data):
    """Generate QR code from data and save it."""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    # Generate a filename based on the current timestamp
    timestamp = int(time.time())
    qr_code_path = f"qr_code_{timestamp}.png"

    # Save the image
    img.save(qr_code_path)
    
    return qr_code_path

def display_qr_code(qr_code_path):
    """Display QR code image."""
    image = Image.open(qr_code_path)
    display(image)

def encrypt_or_decrypt_flow():
    """Main flow for encrypting or decrypting based on user input."""
    choice = input("Would you like to input your own private key? (yes/no): ")
    if choice.lower() == 'yes':
        key, encoded_key = input_private_key()
    else:
        key, encoded_key = generate_private_key()

    data = input("Please enter the text to encrypt: ")
    encrypted_data, nonce, tag = encrypt_data(data, key)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Generate QR code for the private key
    qr_code_path = generate_qr_code(encoded_key)
    
    save_data_to_excel(now, encoded_key, data, encrypted_data, qr_code_path)

    print("Encrypted Data:", encrypted_data)
    print("Encryption Strength:", assess_encryption_strength(encrypted_data))
    print("QR Code:")
    display_qr_code(qr_code_path)

    if input("Do you want to decrypt this data now? (yes/no) ").lower() == 'yes':
        decrypted_data = decrypt_data(encrypted_data, nonce, tag, key)
        print("Decrypted Data:", decrypted_data)
        if decrypted_data == data:
            print("Encryption and decryption is successful.")

if __name__ == '__main__':
    encrypt_or_decrypt_flow()
